import pandas as pd 
import openpyxl
from openpyxl.styles.fonts import Font
from openpyxl.styles import Alignment
from openpyxl.styles.borders import Border, Side

# mergeするセルの範囲を指定
def calculation_range(tag_space, lesson_row, text_book_row, teacher_row, price_row, lesson_code_row, i):
    lesson_space = "A" + str(lesson_row) + ":J" + str(lesson_row+1)
    text_book_space = "A" + str(text_book_row) + ":J" + str(text_book_row+2)
    teacher_space = "B"+str(teacher_row) + ":E" + str(teacher_row+1)
    price_space = "F" + str(price_row) + ":J" + str(price_row+1)
    lesson_code_space = "B" + str(lesson_code_row) + ":E" + str(lesson_code_row+1)
    border_space = "A" + str(tag_space * (i+1)) + ":J" + str(tag_space * (i+1))
    return lesson_space, text_book_space, teacher_space, price_space, lesson_code_space, border_space

# 棚札の作成
def create_text_book_tag(df):
    # excelの準備
    book = openpyxl.load_workbook("./test.xlsx")
    sheet = book["Sheet1"]

    # 入力するセルの行の初期設定
    tag_space = 13
    lesson_row = 2
    text_book_row = 4
    teacher_row = 8
    price_row = 8
    lesson_code_row = 10

    for i in range(len(df)):
        lesson_space, text_book_space, teacher_space, price_space, lesson_code_space, border_space = calculation_range(tag_space, lesson_row, text_book_row, teacher_row, price_row, lesson_code_row, i)

        # セルのmerge
        sheet.merge_cells(lesson_space)
        sheet.merge_cells(text_book_space)
        sheet.merge_cells(teacher_space)
        sheet.merge_cells(price_space)
        sheet.merge_cells(lesson_code_space)

        # セルに入力
        sheet.cell(row=lesson_row,column=1).value = "講義名:" + df.at[i, "講義題目"]
        sheet.cell(row=text_book_row,column=1).value = df.at[i, "教科書名"]
        sheet.cell(row=teacher_row,column=2).value = "担当教員:" + df.at[i, "担当教員"]
        sheet.cell(row=price_row,column=6).value = "値段:" + str(df.at[i, "値段"])
        sheet.cell(row=lesson_code_row,column=2).value = "時間割コード:" + str(df.at[i, "時間割コード"])

        # フォント設定
        sheet.cell(row=lesson_row,column=1).font = Font(name="游ゴシック",size=20,bold=True)
        sheet.cell(row=text_book_row,column=1).font = Font(name="游ゴシック",size=28,bold=True)
        sheet.cell(row=teacher_row,column=2).font = Font(name="游ゴシック",size=14,bold=True)
        sheet.cell(row=price_row,column=6).font = Font(name="游ゴシック",size=14,bold=True)
        sheet.cell(row=lesson_code_row,column=2).font = Font(name="游ゴシック",size=14,bold=True)

        # 中央揃え
        sheet.cell(row=lesson_row,column=1).alignment = Alignment(horizontal="centerContinuous")
        sheet.cell(row=text_book_row,column=1).alignment = Alignment(horizontal="centerContinuous")

        # 罫線を引く
        if i % 3 != 2:
            line = Side(style='medium', color='000000')
            border_bottom = Border(bottom=line)
            # border_cell = "A"+str(tag_space * (i+1))
            # sheet.cell(row=tag_space * (i+1),column=1).border = border_bottom
            for row in sheet[border_space]:
                for cell in row:
                    cell.border = border_bottom

        # 値の更新
        lesson_row += tag_space
        text_book_row += tag_space
        teacher_row += tag_space
        price_row += tag_space
        lesson_code_row += tag_space

    book.save("./save.xlsx")

if __name__ == "__main__":
    df = pd.read_excel("./text_book_data.xlsx")
    create_text_book_tag(df)